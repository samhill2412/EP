import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string
import os
import random


cs_filepath = ""
lib_filepath = ""
dest_folder = ""

def select_dest():
    global dest_folder
    dest_folder = filedialog.askdirectory()
    button_dest.config(text="Destination Folder: " + "\n" + dest_folder, bg="#34a4eb",font=("Arial", 8, "bold"), fg = "white")

def auto_cs():
    button_gen.config(text="Generating Cue Sheet...", bg="#34a4eb",font=("Arial", 12, "bold"), fg = "white")
    error.configure(state="normal")
    error.delete('1.0', tk.END)
    if cs_filepath and lib_filepath and dest_folder:
        if entry_cs.get():
            search_cs = entry_cs.get().upper()
        else:
            search_cs = "B"
        if entry_lib.get():
            search_lib = entry_lib.get().upper()
        else:
            search_lib = "B"

        # Load both workbooks
        cs_wb = load_workbook(cs_filepath)
        lib_wb = load_workbook(lib_filepath)

        # Get the active worksheets
        cs_ws = cs_wb.active
        lib_ws = lib_wb.active

        # Create a new workbook to store appended data
        appended_wb = Workbook()
        appended_ws = appended_wb.active

        # Dictionary to store rows from target file indexed by value from column A
        # Standardize lib_data
        lib_column = column_index_from_string(search_lib)
        print(lib_column)
        lib_data = {str(row[int(lib_column)-1]).lower().replace(" ", "").replace("_", "").replace("(", "").replace(")", ""): row for row in lib_ws.iter_rows(values_only=True)}
        # Iterate through each value in column A of the source worksheet
        for cs_cell in cs_ws[search_cs]:
            cs_value = cs_cell.value
            if cs_value:
                cs_standardised = str(cs_value).lower().replace(" ", "").replace("_", "").replace("(", "").replace(")", "")
                # Check if the value exists in the target data dictionary
                if cs_standardised in lib_data:
                    # Appends source value and the corresponding row from the target worksheet
                    appended_ws.append([cs_value] + list(lib_data[cs_standardised]))
                else:
                    # Appends source value
                    appended_ws.append([cs_value])

        # Determine the directory and filename for the appended workbook
        source_dir, source_filename = os.path.split(cs_filepath)
        appended_filename = os.path.splitext(source_filename)[0] + "_AutoCued.xlsx"
        random_filename = os.path.splitext(source_filename)[0] + str("_AutoCued" + "".join(random.choices('0123456789', k=3)) + ".xlsx")
        appended_path = os.path.join(dest_folder, appended_filename)
        random_path = os.path.join(dest_folder, random_filename)
        if os.path.exists(appended_path):
            appended_wb.save(random_path)
        else:  
            # Save the appended workbook
            appended_wb.save(appended_path)
        button_gen.config(text="Done!", bg="#34a4eb",font=("Arial", 24, "bold"), fg = "white")
    else:
        missing_list = []
        if not cs_filepath:
            missing_list.append("cue sheet")
        if not lib_filepath:
            missing_list.append("library")
        if not dest_folder:
            missing_list.append("destination folder")
        if len(missing_list) == 1:
            missing = missing_list[0] + "."
        if len(missing_list) == 2:
            missing = missing_list[0]  + " and " + missing_list[1] + "."
        if len(missing_list) == 3:
            missing = missing_list[0] + ", " + missing_list[1]  + " and " + missing_list[2] + "."
        button_gen.config(bg="#ff6b70", fg = "white")
        error.insert(tk.END, "Missing: ", "bold")
        error.insert(tk.END, missing, "normal")
        error.configure(state="disabled")
        root.after(200, update_button_text)

def update_button_text():
    button_gen.config(bg="white", fg="black")
    
def select_cs():
    global cs_filepath
    filepath = filedialog.askopenfilename()
    if filepath and ".xls" in filepath:
        cs_filepath = filepath
        button_cs.config(text="Selected Cue Sheet: " + "\n" + cs_filepath.split("/")[-1], bg="#34a4eb",font=("Arial", 8, "bold"), fg = "white")

def select_lib():
    global lib_filepath
    filepath = filedialog.askopenfilename()
    if filepath and ".xls" in filepath:
        lib_filepath = filepath
        button_lib.config(text="Selected Library: " + "\n" + lib_filepath.split("/")[-1], bg="#34a4eb",font=("Arial", 8, "bold"), fg = "white")

root = tk.Tk()
root.title("Sam's Auto Cue Sheet")
root.geometry("720x860")
root.resizable(False, False)
root.configure(bg="light blue")

title = tk.Label(root, text="Auto Cue Sheet", font=("Arial", 24,"bold", "underline"))
title.config(borderwidth=0, highlightthickness=0, bg="light blue")
title.grid(row=1, column=2, columnspan=2, sticky="nsew", padx=0, pady=10)

column_cs = tk.Label(root, text="Column to search:", font=("Arial", 8))
column_cs.config(borderwidth=0, highlightthickness=0, bg="light blue")
column_cs.grid(row=3, column=1, sticky="e")

column_lib = tk.Label(root, text="Column to search:", font=("Arial", 8))
column_lib.config(borderwidth=0, highlightthickness=0, bg="light blue")
column_lib.grid(row=3, column=3, sticky="e")

signature = tk.Text(root, height=3, width=16)
signature.insert(tk.END, "Created by", "underline")
signature.insert(tk.END, "\n Sam Hill", "normal")
signature.tag_config("underline", font=("Arial", 8, "underline"))
signature.tag_config("normal", font=("Arial", 8))
signature.config(borderwidth=0, highlightthickness=0, bg="light blue")
signature.grid(row=1, column=1, padx=0, pady=0)
signature.configure(state="disabled")

contact = tk.Text(root, height=3, width=16)
contact.insert(tk.END, "samuelhill @ \n electricpictures.com.au", "normal")
contact.tag_config("normal", font=("Arial", 8))
contact.config(borderwidth=0, highlightthickness=0, bg="light blue")
contact.grid(row=1, column=4, padx=0, pady=0)
contact.configure(state="disabled")

error = tk.Text(root, height=3, width=16)
error.tag_config("bold", font=("Arial", 8, "bold"))
error.tag_config("normal", font=("Arial", 8))
error.config(borderwidth=0, highlightthickness=0, bg="light blue")
error.configure(state="disabled")
error.grid(row=6, column=2, columnspan=2, sticky="nsew", padx=25, pady=25)

button_cs = tk.Button(root, text="Select Cue Sheet", width=16, height=8, font=("Arial", 12), command=select_cs, padx=20, pady=20, bg="#e0f3ff")
button_cs.grid(row=2, column=1, columnspan=2, sticky="nsew", padx=25, pady=25)

entry_cs = tk.Entry(root, width=3)
entry_cs.grid(row=3, column=2, sticky="w", padx=5, pady=5)
entry_cs.insert(0, "B")

button_lib = tk.Button(root, text="Select Library", width=16, height=8, font=("Arial", 12), command=select_lib, padx=20, pady=20, bg="#e0f3ff")
button_lib.grid(row=2, column=3, columnspan=2, sticky="nsew", padx=25, pady=25)

entry_lib = tk.Entry(root, width=3)
entry_lib.grid(row=3, column=4, sticky="w", padx=5, pady=5)
entry_lib.insert(0, "B")


button_dest = tk.Button(root, text="Select destination folder", width=16, height=8, font=("Arial", 12), command=select_dest, padx=20, pady=20, bg="#e0f3ff")
button_dest.grid(row=4, column=1, columnspan=4, sticky="nsew", padx=25, pady=25)

button_gen = tk.Button(root, text="Generate Cue Sheet", width=16, height=8, font=("Arial", 12), command=auto_cs, padx=20, pady=20, bg="#e0f3ff")
button_gen.grid(row=5, column=1, columnspan=4, sticky="nsew", padx=25, pady=25)

root.grid_rowconfigure(0, weight=1)
root.grid_rowconfigure(5, weight=1)
root.grid_columnconfigure(0, weight=1)
root.grid_columnconfigure(1, weight=1)
root.grid_columnconfigure(2, weight=1)
root.grid_columnconfigure(3, weight=1)
root.grid_columnconfigure(4, weight=1)
root.grid_columnconfigure(5, weight=1)

root.mainloop()