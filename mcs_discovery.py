from openpyxl import load_workbook
from openpyxl import Workbook
import os

def append_matching_data(source_file, target_file):
    # Load both workbooks
    source_wb = load_workbook(source_file)
    target_wb = load_workbook(target_file)

    # Get the active worksheets
    source_ws = source_wb.active
    target_ws = target_wb.active

    # Create a new workbook to store appended data
    appended_wb = Workbook()
    appended_ws = appended_wb.active

    # Dictionary to store rows from target file indexed by value from column A
    # Standardize target_data
    target_data = {str(row[0]).lower().replace(" ", "").replace("_", "").replace("(", "").replace(")", ""): row for row in target_ws.iter_rows(values_only=True)}
    target_data_version = {str(row[0]).lower().replace(" ", "").replace("_", "").replace("(", "").replace(")", "")+str(row[1]).lower().replace(" ", "").replace("_", "").replace("(", "").replace(")", ""): row for row in target_ws.iter_rows(values_only=True)}
    # Iterate through each value in column A of the source worksheet
    for source_cell in source_ws['A']:
        source_value = source_cell.value
        if source_value:
            sv_standardised = str(source_value).lower().replace(" ", "").replace("_", "").replace("(", "").replace(")", "")
            # Check if the value exists in the target data dictionary
            if sv_standardised in target_data:
                # Appends source value and the corresponding row from the target worksheet
                appended_ws.append([source_value] + list(target_data[sv_standardised])) 
            elif sv_standardised in target_data_version:
                appended_ws.append([source_value] + list(target_data_version[sv_standardised])) 
            else:
                # Appends source value
                appended_ws.append([source_value])

    # Determine the directory and filename for the appended workbook
    source_dir, source_filename = os.path.split(source_file)
    appended_filename = os.path.splitext(source_filename)[0] + "_DiscoChecked.xlsx"
    appended_path = os.path.join("V:\Projects\Python Scripts\Music Cue Sheet Script\Completed", appended_filename)

    # Save the appended workbook
    appended_wb.save(appended_path)

if __name__ == "__main__":
    # Replace these with the paths to your actual Excel files
    source_file = "V:\\Projects\\Python Scripts\\Music Cue Sheet Script\\Completed\\agh908 music edl_MusicCueSheet_AllEpsChecked.xlsx"
    target_file = "V:\Projects\Python Scripts\Music Cue Sheet Script\Examples\Discovery Master Database_ESSENTIALS ONLY.xlsx"

    append_matching_data(source_file, target_file)